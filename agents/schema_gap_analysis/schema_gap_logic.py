import pandas as pd
from difflib import SequenceMatcher
from datetime import datetime
from io import BytesIO

from databricks.sdk import WorkspaceClient
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

from agents.schema_gap_analysis.db_schema_loader import load_schema_from_db


# -----------------------------------------------------------------------------
# Databricks client (to read from /Volumes safely)
# -----------------------------------------------------------------------------
ws = WorkspaceClient()


def read_excel_from_volume(path: str, sheet_name=0) -> pd.DataFrame:
    """
    Safely read Excel files stored in Databricks Volumes.
    """
    resp = ws.files.download(path)
    return pd.read_excel(BytesIO(resp.contents.read()), sheet_name=sheet_name)


# =============================================================================
# Helper Functions
# =============================================================================

def similar(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def get_sample_values(df, column, max_values=3):
    if df is None or column not in df.columns:
        return ""
    return "; ".join(df[column].dropna().astype(str).unique()[:max_values])


def get_data_gap_scenario_and_reason(struct_gap, v1, v2):
    if struct_gap == "Missing in New System":
        return ("Column exists in SIMS1 but missing in SIMS2", "Mostly NULL in New system")

    if struct_gap == "Added in New System":
        return ("Column exists only in SIMS2", "Review required")

    if v1 == v2:
        return ("Exact or renamed column with same population", "No Data Gap")

    if v1 and v2:
        return ("Exact structure but values differ", "Business validation required")

    if v1 and not v2:
        return ("Column populated in SIMS1 but NULL in SIMS2", "Mostly NULL in New system")

    if not v1 and v2:
        return ("Column populated in SIMS2 but NULL in SIMS1", "Mostly NULL in Old system")

    return ("", "")


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================

def run_schema_gap_analysis(metadata: dict) -> dict:
    """
    metadata["mode"] = manual | database
    """

    mode = metadata.get("mode")

    # =====================================================
    # ✅ MANUAL MODE – OLD LOGIC, NEW PATHS, VOLUME-SAFE
    # =====================================================
    if mode == "manual":

        paths = metadata["paths"]

        # Required files
        sims1_schema = read_excel_from_volume(paths["sims1_schema"])
        sims2_schema = read_excel_from_volume(paths["sims2_schema"])
        object_map   = read_excel_from_volume(paths["object_mapping"])

        # Optional sample data (sheet_name=None as in old script)
        sims1_data = (
            read_excel_from_volume(paths["sims1_sample"], sheet_name=None)
            if paths.get("sims1_sample")
            else {}
        )

        sims2_data = (
            read_excel_from_volume(paths["sims2_sample"], sheet_name=None)
            if paths.get("sims2_sample")
            else {}
        )

    # =====================================================
    # ✅ DATABASE MODE (UNCHANGED)
    # =====================================================
    elif mode == "database":

        sims1_schema = pd.DataFrame(load_schema_from_db(**metadata["sims1"]))
        sims2_schema = pd.DataFrame(load_schema_from_db(**metadata["sims2"]))
        object_map   = pd.read_excel(metadata["object_mapping_file"])

        sims1_data = {}
        sims2_data = {}

    else:
        raise ValueError("Invalid mode. Expected 'manual' or 'database'.")

    # =====================================================
    # Build Column Description Lookup
    # =====================================================

    sims1_desc_map = dict(
        zip(
            zip(sims1_schema["table_name"], sims1_schema["column_name"]),
            sims1_schema["column_definition"],
        )
    )

    sims2_desc_map = dict(
        zip(
            zip(sims2_schema["table_name"], sims2_schema["column_name"]),
            sims2_schema["column_definition"],
        )
    )

    # =====================================================
    # Strict Mapping & Gap Logic (UNCHANGED)
    # =====================================================

    rows = []

    for _, obj in object_map.iterrows():

        s1_table = obj["sims1_object_name"]
        s2_table = obj["sims2_object_name"]

        s1_cols = sims1_schema[sims1_schema["table_name"] == s1_table]
        s2_cols = sims2_schema[sims2_schema["table_name"] == s2_table]

        s1_df = sims1_data.get(s1_table)
        s2_df = sims2_data.get(s2_table)

        used_s2_cols = set()

        for _, r1 in s1_cols.iterrows():

            s1_col = r1["column_name"]
            s1_dtype = r1["data_type"]
            s1_desc = sims1_desc_map.get((s1_table, s1_col), "")

            exact = s2_cols[
                (s2_cols["column_name"] == s1_col)
                & (~s2_cols["column_name"].isin(used_s2_cols))
            ]

            if not exact.empty:
                r2 = exact.iloc[0]
                s2_col = r2["column_name"]
                s2_dtype = r2["data_type"]
                structure_gap = "Attribute Name Match"
                used_s2_cols.add(s2_col)

            else:
                best_match = None
                best_score = 0

                for _, r2 in s2_cols.iterrows():
                    if r2["column_name"] in used_s2_cols:
                        continue

                    score = similar(s1_col, r2["column_name"])
                    if score >= 0.7 and score > best_score:
                        best_match = r2
                        best_score = score

                if best_match is not None:
                    s2_col = best_match["column_name"]
                    s2_dtype = best_match["data_type"]
                    structure_gap = "Attribute Name Possible Match/Rename"
                    used_s2_cols.add(s2_col)
                else:
                    s2_col = ""
                    s2_dtype = ""
                    structure_gap = "Missing in New System"

            s2_desc = sims2_desc_map.get((s2_table, s2_col), "") if s2_col else ""

            v1 = get_sample_values(s1_df, s1_col)
            v2 = get_sample_values(s2_df, s2_col) if s2_col else ""

            scenario, reason = get_data_gap_scenario_and_reason(
                structure_gap, v1, v2
            )

            rows.append([
                s1_table, s1_col, s1_dtype, s1_desc,
                s2_table, s2_col, s2_dtype, s2_desc,
                structure_gap, v1, v2, scenario, reason
            ])

        for _, r2 in s2_cols.iterrows():
            if r2["column_name"] not in used_s2_cols:
                rows.append([
                    s1_table, "", "", "",
                    s2_table, r2["column_name"], r2["data_type"],
                    sims2_desc_map.get((s2_table, r2["column_name"]), ""),
                    "Added in New System", "",
                    get_sample_values(s2_df, r2["column_name"]),
                    "Column exists only in SIMS2", "Review required"
                ])

    # =====================================================
    # Build Output DataFrame
    # =====================================================

    final_df = pd.DataFrame(rows, columns=[
        "SIMS1 Object Name",
        "SIMS1 Column Name",
        "SIMS1 Column Data Type",
        "SIMS1 Column Description",
        "SIMS2 Object Name",
        "SIMS2 Column Name",
        "SIMS2 Column Data Type",
        "SIMS2 Column Description",
        "Structure Gap Reason",
        "SIMS1 sample value",
        "SIMS2 sample value",
        "Data Gap Scenario",
        "Data Gap Reason"
    ])

    # =====================================================
    # Write Excel to Memory (UI download)
    # =====================================================

    buffer = BytesIO()
    final_df.to_excel(buffer, index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active
    ws.freeze_panes = "A2"

    header_font = Font(bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    GREEN = PatternFill("solid", fgColor="C6EFCE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    RED = PatternFill("solid", fgColor="FFC7CE")
    BLUE = PatternFill("solid", fgColor="BDD7EE")
    PURPLE = PatternFill("solid", fgColor="E4DFEC")

    for cell in ws[1]:
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    STRUCT_COL = 9
    DATA_COL = 13

    for row in ws.iter_rows(min_row=2):
        sg = row[STRUCT_COL - 1].value
        dg = row[DATA_COL - 1].value

        if sg == "Attribute Name Match":
            row[STRUCT_COL - 1].fill = GREEN
        elif sg == "Attribute Name Possible Match/Rename":
            row[STRUCT_COL - 1].fill = YELLOW
        elif sg == "Missing in New System":
            row[STRUCT_COL - 1].fill = RED
        elif sg == "Added in New System":
            row[STRUCT_COL - 1].fill = PURPLE

        if dg == "No Data Gap":
            row[DATA_COL - 1].fill = GREEN
        elif dg and "NULL" in dg:
            row[DATA_COL - 1].fill = BLUE
        elif dg and "Review" in dg:
            row[DATA_COL - 1].fill = PURPLE
        elif dg:
            row[DATA_COL - 1].fill = RED

        for c in row:
            c.border = thin_border
            c.alignment = Alignment(wrap_text=True, vertical="center")

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return {
        "file_name": f"FINAL_GAP_ANALYSIS_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx",
        "file_bytes": out.getvalue(),
    }
